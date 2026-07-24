from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.core.cache import cache
from django.utils import timezone
import logging
from rest_framework.views import APIView
from .services.analytics_service import AnalyticsService
from .serializers import (
    WholesalerStatsSerializer, OrderStatsSerializer,
    RevenueStatsSerializer, ProductStatsSerializer,
    RetailerKPIStatsSerializer,
)
from django.db import models
from catalog.models import Product
from commerce.models import Order, OrderItem
from django.db.models import Q, Sum, Count, Max, ExpressionWrapper, DecimalField
from django.core.paginator import Paginator
from identity.models import User
from datetime import timedelta
from decimal import Decimal
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def wholesaler_dashboard_summary(request):
    """SINGLE ENDPOINT — replaces all 9 wholesaler dashboard calls."""
    user = request.user

    if user.role != 'wholesaler':
        return Response({'status': 'error', 'message': 'Only wholesalers can access'}, status=403)

    cache_key = f"wholesaler_dashboard:{user.id}"
    cached_data = cache.get(cache_key)
    if cached_data:
        return Response({'status': 'success', 'data': cached_data, 'source': 'cache'})

    try:
        from django.db.models import Sum, Count, Avg, F, Q
        from django.utils import timezone
        from datetime import timedelta
        from catalog.models import Product, ProductImage
        from commerce.models import Order, OrderItem

        today = timezone.now().date()
        last_7_days = today - timedelta(days=7)
        last_30_days = today - timedelta(days=30)

        # ---- 1. STATS ----
        total_products = Product.objects.filter(seller=user, seller_type='wholesaler').count()
        total_orders = Order.objects.filter(wholesaler=user).count()
        total_customers = Order.objects.filter(wholesaler=user).values('customer').distinct().count()
        
        revenue_data = Order.objects.filter(
            wholesaler=user,
            status='completed'
        ).aggregate(
            total_revenue=Sum('total_amount'),
            today_revenue=Sum('total_amount', filter=Q(created_at__date=today)),
            week_revenue=Sum('total_amount', filter=Q(created_at__date__gte=last_7_days)),
            month_revenue=Sum('total_amount', filter=Q(created_at__date__gte=last_30_days))
        )
        
        order_stats = {
            'total_orders': total_orders,
            'pending_orders': Order.objects.filter(wholesaler=user, status='pending').count(),
            'processing_orders': Order.objects.filter(wholesaler=user, status='processing').count(),
            'completed_orders': Order.objects.filter(wholesaler=user, status='completed').count(),
            'cancelled_orders': Order.objects.filter(wholesaler=user, status='cancelled').count(),
            'total_revenue': float(revenue_data['total_revenue'] or 0),
            'today_revenue': float(revenue_data['today_revenue'] or 0),
            'week_revenue': float(revenue_data['week_revenue'] or 0),
            'month_revenue': float(revenue_data['month_revenue'] or 0),
            'total_customers': total_customers,
            'total_products': total_products,
        }

        # ---- 2. SALES ANALYTICS ----
        daily_sales = Order.objects.filter(
            wholesaler=user,
            status='completed',
            created_at__date__gte=last_30_days
        ).values('created_at__date').annotate(
            revenue=Sum('total_amount'),
            orders=Count('id')
        ).order_by('created_at__date')

        sales_analytics = {
            'daily': [
                {
                    'date': str(item['created_at__date']),
                    'revenue': float(item['revenue'] or 0),
                    'orders': item['orders']
                }
                for item in daily_sales
            ],
            'total_revenue': float(revenue_data['total_revenue'] or 0),
            'growth': 0
        }

        # ---- 3. CATEGORY PERFORMANCE (FIXED) ----
        category_performance = Product.objects.filter(
            seller=user,
            seller_type='wholesaler'
        ).values('category__name').annotate(
            total_sold=Sum('orderitem__quantity'),
            total_revenue=Sum(F('orderitem__quantity') * F('orderitem__price')),  # ✅ FIXED
            product_count=Count('id')
        ).order_by('-total_revenue')[:10]

        categories_data = [
            {
                'category': item['category__name'] or 'Uncategorized',
                'total_sold': item['total_sold'] or 0,
                'total_revenue': float(item['total_revenue'] or 0),
                'product_count': item['product_count']
            }
            for item in category_performance
        ]

        # ---- 4. RECENT ORDERS ----
        recent_orders = Order.objects.filter(
            wholesaler=user
        ).select_related('customer').order_by('-created_at')[:10]

        recent_orders_data = [
            {
                'id': o.id,
                'order_number': o.order_number or f"ORD-{o.id}",
                'customer': o.customer.get_full_name() or o.customer.email if o.customer else 'Guest',
                'total_amount': float(o.total_amount),
                'status': o.status,
                'created_at': o.created_at.isoformat(),
            }
            for o in recent_orders
        ]

        # ---- 5. LOW STOCK ALERTS ----
        low_stock_products = Product.objects.filter(
            seller=user,
            seller_type='wholesaler',
            status='active',
            stock__lte=F('threshold')
        )[:10]

        low_stock_ids = [p.id for p in low_stock_products]
        low_stock_image_map = {}
        if low_stock_ids:
            images = ProductImage.objects.filter(product_id__in=low_stock_ids, is_primary=True)
            for img in images:
                if img.product_id not in low_stock_image_map:
                    low_stock_image_map[img.product_id] = img.image.url if img.image else None

        low_stock_alerts = [
            {
                'id': p.id,
                'name': p.name,
                'sku': p.sku,
                'currentStock': p.stock,
                'reorderLevel': p.threshold,
                'image_url': low_stock_image_map.get(p.id),
                'status': 'critical' if p.stock == 0 or p.stock <= p.threshold // 2 else 'warning',
            }
            for p in low_stock_products
        ]

        # ---- 6. RECENT ACTIVITY ----
        recent_activities = Order.objects.filter(
            wholesaler=user
        ).select_related('customer').order_by('-created_at')[:8]

        activities_data = [
            {
                'id': o.id,
                'type': 'order',
                'message': f"New order #{o.order_number or o.id} from {o.customer.get_full_name() or 'Guest'}",
                'amount': float(o.total_amount),
                'status': o.status,
                'time': o.created_at.strftime('%I:%M %p'),
                'date': o.created_at.strftime('%b %d, %Y'),
            }
            for o in recent_activities
        ]

        # ---- 7. TOP CUSTOMERS ----
        top_customers = Order.objects.filter(
            wholesaler=user,
            status='completed'
        ).values('customer').annotate(
            total_spent=Sum('total_amount'),
            order_count=Count('id')
        ).order_by('-total_spent')[:6]

        customer_ids = [c['customer'] for c in top_customers if c['customer']]
        customer_map = {u.id: u for u in User.objects.filter(id__in=customer_ids)}

        top_customers_data = [
            {
                'id': c['customer'],
                'name': customer_map.get(c['customer']).get_full_name() or 'Unknown',
                'total_spent': float(c['total_spent'] or 0),
                'order_count': c['order_count'],
            }
            for c in top_customers if c['customer'] in customer_map
        ]

        # ---- 8. PENDING TASKS ----
        pending_orders = Order.objects.filter(
            wholesaler=user,
            status__in=['pending', 'processing']
        ).select_related('customer')[:8]

        pending_tasks = [
            {
                'id': o.id,
                'type': 'order',
                'title': f"Order #{o.order_number or o.id}",
                'description': f"Customer: {o.customer.get_full_name() or 'Guest'}",
                'status': o.status,
                'priority': 'high' if o.status == 'pending' else 'medium',
                'due_date': o.created_at.strftime('%b %d'),
                'created_at': o.created_at.isoformat(),
            }
            for o in pending_orders
        ]

        # ---- 9. QUICK INSIGHTS ----
        quick_insights = {
            'total_revenue': float(revenue_data['total_revenue'] or 0),
            'total_orders': total_orders,
            'total_products': total_products,
            'total_customers': total_customers,
            'low_stock_count': low_stock_products.count(),
            'pending_orders_count': Order.objects.filter(wholesaler=user, status='pending').count(),
            'average_order_value': float(revenue_data['total_revenue'] / total_orders if total_orders > 0 else 0),
            'growth_percentage': 0,
        }

        # ---- FINAL RESPONSE ----
        response_data = {
            'stats': order_stats,
            'salesAnalytics': sales_analytics,
            'categoryPerformance': categories_data,
            'recentOrders': recent_orders_data,
            'lowStockAlerts': low_stock_alerts,
            'recentActivity': activities_data,
            'topCustomers': top_customers_data,
            'pendingTasks': pending_tasks,
            'quickInsights': quick_insights,
        }

        cache.set(cache_key, response_data, timeout=120)
        return Response({'status': 'success', 'data': response_data})

    except Exception as e:
        logger.error(f"Wholesaler dashboard error: {str(e)}", exc_info=True)
        return Response({
            'status': 'error',
            'message': str(e),
            'detail': 'Unable to load dashboard'
        }, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def wholesaler_dashboard_stats(request):
    """Get all dashboard statistics for wholesaler"""
    user = request.user
    
    if user.role != 'wholesaler':
        return Response({'status': 'error', 'message': 'Only wholesalers can access dashboard stats'}, status=403)
    
    try:
        # ✅ Get date parameters from request
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        range_type = request.GET.get('range')
        
        stats = AnalyticsService.get_wholesaler_stats(user, start_date, end_date, range_type)
        serializer = WholesalerStatsSerializer(stats)
        
        return Response({
            'status': 'success',
            'data': serializer.data,
            'cached': cache.has_key(f"wholesaler_stats:{user.id}:{start_date}:{end_date}:{range_type}")
        })
        
    except Exception as e:
        logger.error(f"Error fetching wholesaler stats: {e}")
        return Response({'status': 'error', 'message': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def order_stats(request):
    """Get order statistics (today, week, month, total)"""
    user = request.user
    
    if user.role != 'wholesaler':
        return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
    
    stats = AnalyticsService.get_order_stats(user)
    serializer = OrderStatsSerializer(stats)
    
    return Response({'status': 'success', 'data': serializer.data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def revenue_stats(request):
    """Get revenue statistics (today, week, month, total)"""
    user = request.user
    
    if user.role != 'wholesaler':
        return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
    
    stats = AnalyticsService.get_revenue_stats(user)
    serializer = RevenueStatsSerializer(stats)
    
    return Response({'status': 'success', 'data': serializer.data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def product_stats(request):
    """Get product statistics"""
    user = request.user
    
    if user.role != 'wholesaler':
        return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
    
    stats = AnalyticsService.get_product_stats(user)
    serializer = ProductStatsSerializer(stats)
    
    return Response({'status': 'success', 'data': serializer.data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sales_analytics(request):
    """Get sales analytics chart data (daily/weekly/monthly)"""
    from commerce.models import Order
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import timedelta
    
    user = request.user
    
    if user.role != 'wholesaler':
        return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
    
    period = request.GET.get('period', 'weekly')  # daily, weekly, monthly
    
    today = timezone.now().date()
    
    if period == 'daily':
        days = 7
        dates = [(today - timedelta(days=i)) for i in range(days - 1, -1, -1)]
        labels = [d.strftime('%a') for d in dates]
    elif period == 'weekly':
        weeks = 6
        dates = [(today - timedelta(weeks=i)) for i in range(weeks - 1, -1, -1)]
        labels = [f"Week {d.isocalendar()[1]}" for d in dates]
    else:  # monthly
        months = 6
        dates = [(today.replace(day=1) - timedelta(days=30*i)) for i in range(months - 1, -1, -1)]
        labels = [d.strftime('%b %Y') for d in dates]
    
    sales_data = []
    for date in dates:
        if period == 'daily':
            start_date = date
            end_date = date + timedelta(days=1)
        elif period == 'weekly':
            start_date = date - timedelta(days=date.weekday())
            end_date = start_date + timedelta(days=7)
        else:  # monthly
            start_date = date.replace(day=1)
            next_month = start_date.replace(month=start_date.month + 1) if start_date.month < 12 else start_date.replace(year=start_date.year + 1, month=1)
            end_date = next_month
        
        total = Order.objects.filter(
            items__product__seller=user,  # ← Changed from wholesaler=user
            status='delivered',
            created_at__date__gte=start_date,
            created_at__date__lt=end_date
        ).distinct().aggregate(total=Sum('grand_total'))['total'] or 0
        
        sales_data.append(float(total))
    
    return Response({
        'status': 'success',
        'data': {
            'period': period,
            'labels': labels,
            'values': sales_data,
            'max_value': max(sales_data) if sales_data else 0,
            'total': sum(sales_data)
        }
    })


class CategoryPerformanceAPIView(APIView):
    """Get sales performance by category"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Handle GET request"""
        from commerce.models import Order, OrderItem
        from catalog.models import Category
        from django.db.models import Sum
        
        user = request.user
        
        if user.role != 'wholesaler':
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get all order items from wholesaler's delivered orders
        order_items = OrderItem.objects.filter(
            order__items__product__seller=user,  # ← Changed from order__wholesaler=user
            order__status='delivered'
        ).distinct()
        
        # Aggregate by category
        category_data = {}
        for item in order_items:
            category = item.product.category
            if category:
                cat_name = category.name
                if cat_name not in category_data:
                    category_data[cat_name] = {
                        'total': 0,
                        'revenue': 0
                    }
                category_data[cat_name]['total'] += 1
                category_data[cat_name]['revenue'] += float(item.price * item.quantity)
        
        # Calculate percentages and prepare response
        total_revenue = sum(data['revenue'] for data in category_data.values())
        total_items = sum(data['total'] for data in category_data.values())
        
        result = []
        colors = ['primary', 'success', 'accent', 'warning', 'info']
        for idx, (name, data) in enumerate(category_data.items()):
            percentage = (data['revenue'] / total_revenue * 100) if total_revenue > 0 else 0
            result.append({
                'id': idx + 1,
                'name': name,
                'value': round(percentage, 1),
                'color': colors[idx % len(colors)],
                'amount': f"₹{data['revenue']:,.0f}",
                'revenue': data['revenue'],
                'item_count': data['total']
            })
        
        # Sort by revenue descending
        result.sort(key=lambda x: x['revenue'], reverse=True)
        
        return Response({
            'status': 'success',
            'data': {
                'categories': result,
                'total_revenue': f"₹{total_revenue:,.0f}",
                'total_items': total_items
            }
        })
    

class LowStockAlertsAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        if user.role != 'wholesaler':
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
        
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 8))
        
        products = Product.objects.filter(
            seller=user,
            stock__lte=models.F('threshold'),
            stock__gt=0,
            status='active'
        ).select_related('category').order_by('stock')
        
        paginator = Paginator(products, per_page)
        page_obj = paginator.get_page(page)
        
        data = []
        for product in page_obj:
            data.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'stock': product.stock,
                'threshold': product.threshold,
                'category': product.category.name if product.category else 'Uncategorized',
            })
        
        return Response({
            'status': 'success',
            'data': data,
            'count': paginator.count,
            'page': page,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        })
    

class RecentOrdersAPIView(APIView):
    """Get recent orders for wholesaler dashboard"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        if user.role != 'wholesaler':
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
        
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        orders = Order.objects.filter(
            items__product__seller=user  # ← Changed from wholesaler=user
        ).distinct().select_related('customer').order_by('-created_at')
        
        paginator = Paginator(orders, per_page)
        page_obj = paginator.get_page(page)
        
        data = []
        for order in page_obj:
            data.append({
                'id': order.order_number,
                'customer': order.customer.get_full_name() or order.customer.email,
                'items': order.items.count(),
                'amount': float(order.grand_total),
                'date': self.get_time_ago(order.created_at),
                'status': order.status,
                'payment': order.payment_status
            })
        
        return Response({
            'status': 'success',
            'data': data,
            'count': paginator.count,
            'page': page,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        })
    
    def get_time_ago(self, created_at):
        from django.utils import timezone
        from datetime import datetime
        
        now = timezone.now()
        diff = now - created_at
        
        if diff.seconds < 60:
            return f"{diff.seconds} sec ago"
        elif diff.seconds < 3600:
            minutes = diff.seconds // 60
            return f"{minutes} min ago"
        elif diff.days == 0:
            hours = diff.seconds // 3600
            return f"{hours} hour ago"
        elif diff.days == 1:
            return "Yesterday"
        else:
            return f"{diff.days} days ago"
        

class RecentActivityAPIView(APIView):
    """Get recent activity for wholesaler dashboard"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        if user.role != 'wholesaler':
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
        
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 8))
        
        activities = []
        
        # Get recent orders
        recent_orders = Order.objects.filter(items__product__seller=user).distinct().order_by('-created_at')[:20]
        for order in recent_orders:
            activities.append({
                'id': f"order_{order.id}",
                'type': 'order',
                'message': f"Order {order.order_number}",
                'time': self.get_time_ago(order.created_at),
                'amount': float(order.grand_total),
                'status': order.status,
                'icon': 'Package',
                'color': 'primary'
            })
        
        # Sort by time
        activities.sort(key=lambda x: x['time'], reverse=True)
        
        # Paginate
        paginator = Paginator(activities, per_page)
        page_obj = paginator.get_page(page)
        
        return Response({
            'status': 'success',
            'data': page_obj.object_list,
            'count': paginator.count,
            'page': page,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        })
    
    def get_time_ago(self, created_at):
        now = timezone.now()
        diff = now - created_at
        
        if diff.seconds < 60:
            return f"{diff.seconds} sec ago"
        elif diff.seconds < 3600:
            minutes = diff.seconds // 60
            return f"{minutes} min ago"
        elif diff.days == 0:
            hours = diff.seconds // 3600
            return f"{hours} hour ago"
        elif diff.days == 1:
            return "Yesterday"
        else:
            return f"{diff.days} days ago"
        


class TopCustomersAPIView(APIView):
    """Get top customers for wholesaler dashboard"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        if user.role != 'wholesaler':
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
        
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 6))
        
        # Get all retailers who bought from this wholesaler
        customers_data = {}
        orders = Order.objects.filter(
            items__product__seller=user,  # ← Changed from wholesaler=user
            status='delivered'
        ).distinct().select_related('customer')
        
        for order in orders:
            customer = order.customer
            if customer.id not in customers_data:
                customers_data[customer.id] = {
                    'id': customer.id,
                    'name': customer.get_full_name() or customer.email,
                    'email': customer.email,
                    'phone': customer.mobile or '',
                    'orders': 0,
                    'spent': 0,
                    'since': customer.date_joined.year
                }
            customers_data[customer.id]['orders'] += 1
            customers_data[customer.id]['spent'] += float(order.grand_total)
        
        # Convert to list and sort by spent
        customers_list = list(customers_data.values())
        customers_list.sort(key=lambda x: x['spent'], reverse=True)
        
        # Add rank and avatar
        for idx, cust in enumerate(customers_list):
            cust['rank'] = idx + 1
            cust['avatar'] = cust['name'][:2].upper()
            cust['color'] = ['primary', 'success', 'accent', 'warning', 'info'][idx % 5]
            cust['type'] = 'retailer'
            cust['spent_formatted'] = f"₹{cust['spent']:,.0f}"
        
        # Paginate
        paginator = Paginator(customers_list, per_page)
        page_obj = paginator.get_page(page)
        
        # Calculate total growth (mock for now)
        total_spent = sum(c['spent'] for c in customers_list)
        previous_spent = total_spent * 0.77  # Mock 23% growth
        
        return Response({
            'status': 'success',
            'data': page_obj.object_list,
            'count': paginator.count,
            'page': page,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'total_spent': f"₹{total_spent:,.0f}",
            'growth': '23%'
        })
    

class PendingTasksAPIView(APIView):
    """Get pending tasks for wholesaler dashboard"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        if user.role != 'wholesaler':
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
        
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 8))
        task_type = request.GET.get('type', 'all')  # all, orders, products, payouts
        
        tasks = []
        
        # Pending Orders
        if task_type in ['all', 'orders']:
            pending_orders = Order.objects.filter(
                items__product__seller=user,
                status__in=['pending', 'confirmed']
            ).order_by('-created_at')[:20]
            
            for order in pending_orders:
                tasks.append({
                    'id': f"order_{order.id}",
                    'type': 'order',
                    'task': f"Process order {order.order_number}",
                    'time': self.get_time_ago(order.created_at),
                    'priority': self.get_priority(order.created_at),
                    'amount': f"₹{float(order.grand_total):,.0f}",
                    'customer': order.customer.get_full_name() or order.customer.email
                })
        
        # Low Stock Products
        if task_type in ['all', 'products']:
            low_stock_products = Product.objects.filter(
                seller=user,
                stock__lte=models.F('threshold'),
                stock__gt=0,
                status='active'
            ).order_by('stock')[:20]
            
            for product in low_stock_products:
                tasks.append({
                    'id': f"product_{product.id}",
                    'type': 'product',
                    'task': f"Restock {product.name}",
                    'time': f"Stock: {product.stock}/{product.threshold}",
                    'priority': 'high' if product.stock <= 2 else 'medium',
                    'product': product.name,
                    'sku': product.sku
                })
        
        # Sort by priority
        priority_order = {'high': 0, 'medium': 1, 'low': 2}
        tasks.sort(key=lambda x: priority_order.get(x.get('priority', 'low'), 3))
        
        # Calculate stats
        high_count = sum(1 for t in tasks if t.get('priority') == 'high')
        medium_count = sum(1 for t in tasks if t.get('priority') == 'medium')
        low_count = sum(1 for t in tasks if t.get('priority') == 'low')
        
        # Paginate
        paginator = Paginator(tasks, per_page)
        page_obj = paginator.get_page(page)
        
        return Response({
            'status': 'success',
            'data': page_obj.object_list,
            'count': paginator.count,
            'page': page,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'stats': {
                'high': high_count,
                'medium': medium_count,
                'low': low_count,
                'total': len(tasks)
            }
        })
    
    def get_time_ago(self, created_at):
        now = timezone.now()
        diff = now - created_at
        
        if diff.seconds < 3600:
            minutes = diff.seconds // 60
            return f"{minutes} min ago" if minutes > 0 else "Just now"
        elif diff.days == 0:
            hours = diff.seconds // 3600
            return f"{hours} hour ago"
        elif diff.days == 1:
            return "Yesterday"
        else:
            return f"{diff.days} days ago"
    
    def get_priority(self, created_at):
        now = timezone.now()
        diff = now - created_at
        
        if diff.seconds < 3600:  # Less than 1 hour
            return 'high'
        elif diff.days == 0:  # Less than 24 hours
            return 'medium'
        else:
            return 'low'
        

class WithdrawalStatsAPIView(APIView):
    """Get withdrawal statistics for wholesaler"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        if user.role != 'wholesaler':
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=403)
        
        # Get total withdrawals (from your withdrawal model - create if needed)
        # For now, calculate from orders or use mock data
        total_withdrawn = Decimal('0')
        pending_withdrawals = Decimal('0')
        next_payout_date = "Mar 25, 2024"
        
        # Calculate available balance (total_revenue - total_withdrawn)
        from commerce.models import Order
        total_revenue = Order.objects.filter(
            items__product__seller=user,
            status='delivered'
        ).distinct().aggregate(total=models.Sum('grand_total'))['total'] or Decimal('0')
        
        available_balance = total_revenue - total_withdrawn
        
        return Response({
            'status': 'success',
            'data': {
                'total_withdrawn': float(total_withdrawn),
                'pending_withdrawals': float(pending_withdrawals),
                'available_balance': float(available_balance),
                'next_payout_date': next_payout_date,
                'total_revenue': float(total_revenue)
            }
        })
    

class TopProductsAPIView(APIView):
    """Get top selling products for wholesaler"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        if user.role != 'wholesaler':
            return Response({'error': 'Unauthorized'}, status=403)
        
        # Get delivered order items
        top_products = OrderItem.objects.filter(
            order__items__product__seller=user,
            order__status='delivered'
        ).distinct().values('product_id', 'product_name').annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum('total')
        ).order_by('-total_sold')[:10]
        
        return Response({
            'status': 'success',
            'data': list(top_products)
        })


class GeographicSalesAPIView(APIView):
    """Get sales by region/city"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        if user.role != 'wholesaler':
            return Response({'error': 'Unauthorized'}, status=403)
        
        geo_sales = Order.objects.filter(
            items__product__seller=user,
            status='delivered'
        ).distinct().values('shipping_city').annotate(
            total=Sum('grand_total'),
            orders=Count('id')
        ).order_by('-total')[:10]
        
        return Response({
            'status': 'success',
            'data': list(geo_sales)
        })


class HourlySalesAPIView(APIView):
    """Get sales by hour of day"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        if user.role != 'wholesaler':
            return Response({'error': 'Unauthorized'}, status=403)
        
        # Get orders from last 30 days
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        hourly_data = []
        for hour in range(24):
            total = Order.objects.filter(
                items__product__seller=user,
                status='delivered',
                created_at__hour=hour,
                created_at__gte=thirty_days_ago
            ).distinct().aggregate(total=Sum('grand_total'))['total'] or 0
            
            hourly_data.append({
                'hour': f"{hour}:00",
                'total': float(total)
            })
        
        return Response({
            'status': 'success',
            'data': hourly_data
        })
    

class ExportReportAPIView(APIView):
    """Export report to PDF/Excel/CSV"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        report_type = request.GET.get('type', 'sales')
        format_type = request.GET.get('format', 'pdf')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        user = request.user
        if user.role != 'wholesaler':
            return Response({'error': 'Unauthorized'}, status=403)
        
        # Get data based on report type
        data = self.get_report_data(user, report_type, start_date, end_date)
        
        if format_type == 'csv':
            return self.export_csv(data, report_type)
        elif format_type == 'excel':
            return self.export_excel(data, report_type)
        else:
            return self.export_pdf(data, report_type)
    
    def get_report_data(self, user, report_type, start_date, end_date):
        from commerce.models import Order, OrderItem
        from catalog.models import Product
        
        # Apply date filter
        date_filter = {}
        if start_date and end_date:
            date_filter = {'created_at__date__range': [start_date, end_date]}
        
        if report_type == 'sales':
            orders = Order.objects.filter(items__product__seller=user, **date_filter).distinct()
            return [{
                'Date': o.created_at.strftime('%Y-%m-%d'),
                'Order ID': o.order_number,
                'Amount': float(o.grand_total),
                'Status': o.status,
                'Payment': o.payment_status
            } for o in orders]
        
        elif report_type == 'inventory':
            products = Product.objects.filter(seller=user)
            return [{
                'SKU': p.sku,
                'Product': p.name,
                'Category': p.category.name if p.category else '-',
                'Stock': p.stock,
                'Threshold': p.threshold,
                'Status': 'Low Stock' if p.stock <= p.threshold else 'In Stock'
            } for p in products]
        
        elif report_type == 'customer':
            customers = Order.objects.filter(items__product__seller=user).distinct().values('customer__email').distinct()
            return [{'Customer': c['customer__email'], 'Orders': 0} for c in customers]
        
        return []
    
    def export_csv(self, data, report_type):
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys() if data else [])
        writer.writeheader()
        writer.writerows(data)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.csv"'
        return response
    
    def export_excel(self, data, report_type):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report_type.capitalize()} Report"
        
        # Headers
        headers = list(data[0].keys()) if data else []
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3662d9", end_color="3662d9", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key, ''))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
        wb.save(response)
        return response
    
    def export_pdf(self, data, report_type):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        elements = []
        elements.append(Paragraph(f"{report_type.capitalize()} Report", styles['Title']))
        
        if data:
            table_data = [list(data[0].keys())] + [[row.get(k, '') for k in data[0].keys()] for row in data]
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.pdf"'
        return response
    



# -------------------------------------------------------------------RETAILERS-------------------------------------------


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_dashboard_summary(request):
    """SINGLE ENDPOINT — replaces all 8 retailer dashboard calls. Bulk queries only."""
    user = request.user

    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)

    cache_key = f"retailer_dashboard:{user.id}"
    cached_data = cache.get(cache_key)
    if cached_data:
        return Response({'status': 'success', 'data': cached_data, 'source': 'cache'})

    try:
        from django.db.models.functions import ExtractHour
        from django.db.models import Sum, Count, Max, F
        from django.utils import timezone
        from datetime import timedelta
        from catalog.models import ProductImage

        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        last_week = today - timedelta(days=7)
        last_30_days = today - timedelta(days=30)
        start_of_day = timezone.datetime.combine(today, timezone.datetime.min.time())
        end_of_day = start_of_day + timedelta(days=1)

        all_orders = Order.objects.filter(retailer=user)
        
        # ---- 1. KPI STATS ----
        today_orders = all_orders.filter(status='completed', created_at__gte=start_of_day, created_at__lt=end_of_day)
        today_sales = today_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        yesterday_sales = all_orders.filter(status='completed', created_at__date=yesterday).aggregate(total=Sum('total_amount'))['total'] or 0
        sales_change = round(((today_sales - yesterday_sales) / yesterday_sales) * 100, 1) if yesterday_sales else (100 if today_sales else 0)
        total_orders_7d = all_orders.filter(created_at__date__gte=last_week).count()
        total_customers = all_orders.values('customer').distinct().count()
        total_products = Product.objects.filter(seller=user, seller_type='retailer').count()

        kpi_stats = {
            'today_sales': {'value': today_sales, 'change': sales_change, 'trend': 'up' if sales_change >= 0 else 'down'},
            'total_orders': {'value': total_orders_7d},
            'total_customers': {'value': total_customers},
            'total_products': {'value': total_products},
        }

        # ---- 2. DAILY SALES / HOURLY BREAKDOWN ----
        hourly_qs = today_orders.filter(created_at__hour__gte=10, created_at__hour__lt=22).annotate(
            hour=ExtractHour('created_at')
        ).values('hour').annotate(
            total=Sum('total_amount'), count=Count('id')
        ).order_by('hour')
        hourly_map = {row['hour']: {'sales': row['total'] or 0, 'transactions': row['count']} for row in hourly_qs}
        daily_sales = [
            {'hour': f"{h} AM" if h < 12 else f"{h-12} PM" if h > 12 else "12 PM",
             'amount': hourly_map.get(h, {}).get('sales', 0),
             'transactions': hourly_map.get(h, {}).get('transactions', 0)}
            for h in range(10, 22)
        ]

        busiest = max(daily_sales, key=lambda h: h['amount']) if daily_sales else {'hour': 'N/A', 'amount': 0, 'transactions': 0}

        # ---- 3. TODAY SUMMARY ----
        total_transactions = today_orders.count()
        average_bill = today_sales / total_transactions if total_transactions else 0
        total_items = today_orders.aggregate(total=Sum('items__quantity'))['total'] or 0
        payment_methods = {
            'upi': today_orders.filter(payment_method='upi').count(),
            'card': today_orders.filter(payment_method='card').count(),
            'cash': today_orders.filter(payment_method='cash').count(),
            'wallet': today_orders.filter(payment_method='wallet').count(),
        }
        avg_daily_target = all_orders.filter(
            status='completed', created_at__date__gte=last_week
        ).aggregate(avg=Sum('total_amount'))['avg'] or 25000
        avg_daily_target = avg_daily_target / 7 if avg_daily_target else 25000

        today_summary = {
            'totalTransactions': total_transactions,
            'averageBill': round(average_bill, 2),
            'totalItems': total_items,
            'revenue': today_sales,
            'target': round(avg_daily_target, 2),
            'paymentMethods': payment_methods,
            'hourlyBreakdown': daily_sales,
            'busiestHour': busiest['hour'],
            'busiestHourSales': busiest['amount'],
            'peakHourCustomers': busiest['transactions'],
        }

        # ---- 4. TOP PRODUCTS (WITH IMAGES) ----
        top_products_qs = list(OrderItem.objects.filter(
            order__retailer=user, order__status='completed', order__created_at__date__gte=last_30_days
        ).values('product_id', 'product__name', 'product__sku', 'product__stock').annotate(
            total_sales=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('price'))
        ).order_by('-total_sales')[:5])

        product_ids = [p['product_id'] for p in top_products_qs if p['product_id']]
        image_map = {}
        if product_ids:
            from catalog.models import ProductImage
            images = ProductImage.objects.filter(product_id__in=product_ids, is_primary=True)
            for img in images:
                if img.product_id not in image_map:
                    image_map[img.product_id] = img.image.url if img.image else None

        top_products = []
        for p in top_products_qs:
            top_products.append({
                'id': p['product_id'],
                'name': p['product__name'],
                'sku': p['product__sku'],
                'sales': p['total_sales'] or 0,
                'revenue': float(p['total_revenue']) if p['total_revenue'] else 0,
                'stock': p['product__stock'] or 0,
                'image_url': image_map.get(p['product_id']),  # ✅ ONLY image_url
            })

        # ---- 5. RECENT TRANSACTIONS ----
        recent_orders = all_orders.select_related('customer').order_by('-created_at')[:10]
        recent_transactions = [{
            'id': f"#TR-{o.id:04d}",
            'customer': o.customer.get_full_name() or o.customer.email.split('@')[0] if o.customer else 'Guest',
            'amount': float(o.total_amount),
            'status': o.status,
            'time': o.created_at.strftime('%I:%M %p'),
        } for o in recent_orders]


        # ---- 6. LOW STOCK ALERTS (WITH IMAGES) ----
        low_stock_qs = Product.objects.filter(
            seller=user, seller_type='retailer', status='active', stock__lte=F('threshold')
        )[:10]

        low_stock_product_ids = [p.id for p in low_stock_qs]
        low_stock_image_map = {}
        if low_stock_product_ids:
            from catalog.models import ProductImage
            low_stock_images = ProductImage.objects.filter(product_id__in=low_stock_product_ids, is_primary=True)
            for img in low_stock_images:
                if img.product_id not in low_stock_image_map:
                    low_stock_image_map[img.product_id] = img.image.url if img.image else None

        low_stock_alerts = [{
            'id': p.id,
            'name': p.name,
            'sku': p.sku,
            'currentStock': p.stock,
            'reorderLevel': p.threshold,
            'image_url': low_stock_image_map.get(p.id),  # ✅ ONLY image_url
            'status': 'critical' if p.stock == 0 or p.stock <= p.threshold // 2 else 'warning',
        } for p in low_stock_qs]

        # ---- 7. CUSTOMER ACTIVITY ----
        customer_qs = all_orders.filter(status='completed').values('customer').annotate(
            total_orders=Count('id'), total_amount=Sum('total_amount'), last_order=Max('created_at')
        ).order_by('-last_order')[:10]
        customer_ids = [c['customer'] for c in customer_qs if c['customer']]
        customers_map = {c.id: c for c in User.objects.filter(id__in=customer_ids)}
        customer_activity = []
        for c in customer_qs:
            cust = customers_map.get(c['customer'])
            if not cust:
                continue
            customer_activity.append({
                'id': cust.id,
                'name': cust.get_full_name() or cust.email.split('@')[0],
                'visits': c['total_orders'],
                'amount': c['total_amount'] or 0,
                'lastVisit': c['last_order'].strftime('%b %d, %I:%M %p') if c['last_order'] else 'Never',
                'status': 'vip' if c['total_orders'] > 10 else 'active',
            })

        # ---- 8. QUICK REORDER (WITH IMAGES) ----
        products_qs = Product.objects.filter(seller=user, seller_type='retailer', status='active')

        # ✅ Pre-fetch images
        all_product_ids = [p.id for p in products_qs]
        reorder_image_map = {}
        if all_product_ids:
            from catalog.models import ProductImage
            images = ProductImage.objects.filter(product_id__in=all_product_ids, is_primary=True)
            for img in images:
                if img.product_id not in reorder_image_map:
                    reorder_image_map[img.product_id] = img.image.url if img.image else None

        sales_map_qs = OrderItem.objects.filter(
            order__retailer=user, order__status='completed', order__created_at__date__gte=last_30_days,
            product__in=products_qs
        ).values('product_id').annotate(total_sold=Sum('quantity'))
        sales_map = {row['product_id']: row['total_sold'] or 0 for row in sales_map_qs}

        reorder_suggestions = []
        for p in products_qs:
            total_sold = sales_map.get(p.id, 0)
            daily_velocity = total_sold / 30 if total_sold else 0
            days_until_out = int(p.stock / daily_velocity) if daily_velocity > 0 and p.stock > 0 else 999
            needs_reorder = p.stock <= p.threshold or (0 < days_until_out < 7)
            if not needs_reorder:
                continue
            urgency = 'critical' if (p.stock == 0 or days_until_out <= 3) else 'high' if (p.stock <= p.threshold // 2 or days_until_out <= 7) else 'medium'
            suggested_qty = int((daily_velocity * 30) + p.threshold) if daily_velocity > 0 else p.threshold * 2
            reorder_suggestions.append({
                'id': p.id,
                'name': p.name,
                'sku': p.sku,
                'currentStock': p.stock,
                'reorderLevel': p.threshold,
                'salesVelocity': round(daily_velocity, 1),
                'daysUntilOut': days_until_out,
                'suggestedQty': suggested_qty,
                'urgency': urgency,
                'image_url': reorder_image_map.get(p.id),  # ✅ ONLY image_url
            })
        urgency_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        reorder_suggestions.sort(key=lambda x: urgency_order.get(x['urgency'], 4))
        reorder_suggestions = reorder_suggestions[:10]

        # ---- FINAL RESPONSE ----
        response_data = {
            'kpiStats': kpi_stats,
            'todaySummary': today_summary,
            'dailySales': daily_sales,
            'topProducts': top_products,
            'recentTransactions': recent_transactions,
            'lowStockAlerts': low_stock_alerts,
            'customerActivity': customer_activity,
            'quickReorder': reorder_suggestions,
        }

        cache.set(cache_key, response_data, timeout=120)
        return Response({'status': 'success', 'data': response_data})

    except Exception as e:
        logger.error(f"Retailer dashboard error: {str(e)}", exc_info=True)
        return Response({
            'status': 'error', 
            'message': str(e),
            'detail': 'Unable to load dashboard'
        }, status=500)

    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_kpi_stats(request):
    """Get KPI statistics for retailer dashboard cards"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum, Count
        from django.utils import timezone
        from datetime import timedelta
        
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        last_week = today - timedelta(days=7)
        
        # Today's Sales
        today_sales = Order.objects.filter(
            retailer=user,
            created_at__date=today,
            status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        yesterday_sales = Order.objects.filter(
            retailer=user,
            created_at__date=yesterday,
            status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        if yesterday_sales > 0:
            sales_change = round(((today_sales - yesterday_sales) / yesterday_sales) * 100, 1)
        else:
            sales_change = 100 if today_sales > 0 else 0
        
        # Total Orders (last 7 days)
        total_orders = Order.objects.filter(
            retailer=user,
            created_at__date__gte=last_week
        ).count()
        
        prev_week_orders = Order.objects.filter(
            retailer=user,
            created_at__date__range=[last_week - timedelta(days=7), last_week]
        ).count()
        
        if prev_week_orders > 0:
            orders_change = round(((total_orders - prev_week_orders) / prev_week_orders) * 100, 1)
        else:
            orders_change = 100 if total_orders > 0 else 0
        
        # Total Customers
        total_customers = Order.objects.filter(retailer=user).values('customer').distinct().count()
        
        # Total Products
        total_products = Product.objects.filter(
            seller=user,
            seller_type='retailer'
        ).count()
        
        stats = {
            'today_sales': {
                'value': today_sales,
                'change': sales_change,
                'trend': 'up' if sales_change >= 0 else 'down',
                'period': 'vs yesterday'
            },
            'total_orders': {
                'value': total_orders,
                'change': orders_change,
                'trend': 'up' if orders_change >= 0 else 'down',
                'period': 'vs last week'
            },
            'total_customers': {
                'value': total_customers,
                'change': 0,
                'trend': 'up',
                'period': 'all time'
            },
            'total_products': {
                'value': total_products,
                'change': 0,
                'trend': 'up',
                'period': 'total inventory'
            }
        }
        
        serializer = RetailerKPIStatsSerializer(stats)
        
        return Response({
            'status': 'success',
            'data': serializer.data
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_daily_sales(request):
    """Get hourly sales data for retailer dashboard chart"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum
        from django.utils import timezone
        from datetime import datetime, timedelta
        
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        
        # Get today's hourly sales
        today_sales_data = []
        for hour in range(10, 22):  # 10 AM to 9 PM
            start_time = datetime.combine(today, datetime.min.time()) + timedelta(hours=hour)
            end_time = start_time + timedelta(hours=1)
            
            sales = Order.objects.filter(
                retailer=user,
                status='completed',
                created_at__gte=start_time,
                created_at__lt=end_time
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            today_sales_data.append({
                'hour': f"{hour} AM" if hour < 12 else f"{hour-12} PM" if hour > 12 else "12 PM",
                'sales': sales,
                'target': 0  # Target can be calculated based on average
            })
        
        # Get yesterday's hourly sales
        yesterday_sales_data = []
        for hour in range(10, 22):
            start_time = datetime.combine(yesterday, datetime.min.time()) + timedelta(hours=hour)
            end_time = start_time + timedelta(hours=1)
            
            sales = Order.objects.filter(
                retailer=user,
                status='completed',
                created_at__gte=start_time,
                created_at__lt=end_time
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            yesterday_sales_data.append({
                'hour': f"{hour} AM" if hour < 12 else f"{hour-12} PM" if hour > 12 else "12 PM",
                'sales': sales,
                'target': 0
            })
        
        # Calculate peak hour
        peak_sales = max(today_sales_data, key=lambda x: x['sales']) if today_sales_data else {}
        
        # Calculate total sales
        total_today = sum(item['sales'] for item in today_sales_data)
        total_yesterday = sum(item['sales'] for item in yesterday_sales_data)
        
        # Calculate growth
        growth = round(((total_today - total_yesterday) / total_yesterday) * 100, 1) if total_yesterday > 0 else 0
        
        return Response({
            'status': 'success',
            'data': {
                'today': today_sales_data,
                'yesterday': yesterday_sales_data,
                'peak_hour': {
                    'hour': peak_sales.get('hour', 'N/A'),
                    'sales': peak_sales.get('sales', 0)
                },
                'total_today': total_today,
                'total_yesterday': total_yesterday,
                'growth': growth
            }
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_top_products(request):
    """Get top selling products for retailer dashboard"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum, Count, F
        from django.utils import timezone
        from datetime import timedelta
        
        # Get last 30 days data
        last_30_days = timezone.now().date() - timedelta(days=30)
        
        # Get top selling products
        top_products = OrderItem.objects.filter(
            order__retailer=user,
            order__status='completed',
            order__created_at__date__gte=last_30_days
        ).values(
            'product_id', 'product__name', 'product__sku', 'product__stock'
        ).annotate(
            total_sales=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('price'))
        ).order_by('-total_sales')[:5]
        
        # Calculate previous period for trend
        prev_period_start = last_30_days - timedelta(days=30)
        prev_period_end = last_30_days
        
        products_data = []
        for product in top_products:
            # Get previous period sales for trend
            prev_sales = OrderItem.objects.filter(
                order__retailer=user,
                order__status='completed',
                product_id=product['product_id'],
                order__created_at__date__range=[prev_period_start, prev_period_end]
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            current_sales = product['total_sales'] or 0
            
            if prev_sales > 0:
                trend_percent = round(((current_sales - prev_sales) / prev_sales) * 100, 1)
                trend = f"{'+' if trend_percent >= 0 else ''}{trend_percent}%"
            else:
                trend = '+100%' if current_sales > 0 else '0%'
            
            products_data.append({
                'id': product['product_id'],
                'name': product['product__name'],
                'sku': product['product__sku'],
                'sales': current_sales,
                'revenue': product['total_revenue'] or 0,
                'stock': product['product__stock'] or 0,
                'trend': trend,
                'image': Product.objects.get(id=product['product_id']).images.filter(is_primary=True).first().image.url if Product.objects.get(id=product['product_id']).images.filter(is_primary=True).first() else '📦'
            })
        
        return Response({
            'status': 'success',
            'data': products_data
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_customer_activity(request):
    """Get customer activity for retailer dashboard"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum, Count, Q
        from django.utils import timezone
        from datetime import timedelta
        
        # Get filter parameter
        time_filter = request.GET.get('filter', 'all')  # all, today, repeat, new
        
        customer_ids = Order.objects.filter(
            retailer=user
        ).values_list('customer', flat=True).distinct()
        
        # Base queryset - customers who placed orders with this retailer
        customers = User.objects.filter(
            id__in=customer_ids,
            role='customer',
            is_active=True
        )
        
        # Apply time filter
        today = timezone.now().date()
        
        if time_filter == 'today':
            customers = customers.filter(last_login__date=today)
        elif time_filter == 'repeat':
            # Customers with more than 5 orders
            customers = customers.annotate(
                order_count=Count('orders', filter=Q(orders__status='completed'))
            ).filter(order_count__gt=5)
        elif time_filter == 'new':
            # Customers created in last 7 days
            last_week = today - timedelta(days=7)
            customers = customers.filter(date_joined__date__gte=last_week)
        
        # Get customer data
        customer_data = []
        for customer in customers[:10]:  # Limit to 10 customers
            # Get order statistics
            order_stats = Order.objects.filter(
                customer=customer,
                retailer=user,
                status='completed'
            ).aggregate(
                total_orders=Count('id'),
                total_amount=Sum('total_amount'),
                last_order=Max('created_at')
            )
            
            total_orders = order_stats['total_orders'] or 0
            total_amount = order_stats['total_amount'] or 0
            last_order = order_stats['last_order']
            
            # Determine customer type
            if total_orders > 5:
                customer_type = 'repeat'
            elif total_orders == 1:
                customer_type = 'new'
            else:
                customer_type = 'regular'
            
            # Determine status
            if total_orders > 10:
                status = 'vip'
            elif total_orders > 0:
                status = 'active'
            else:
                status = 'inactive'
            
            # Get initials for avatar
            name_parts = customer.get_full_name().split()
            initials = ''.join([p[0].upper() for p in name_parts[:2]]) if name_parts else customer.email[:2].upper()
            
            customer_data.append({
                'id': customer.id,
                'name': customer.get_full_name() or customer.email.split('@')[0],
                'type': customer_type,
                'visits': total_orders,
                'lastVisit': last_order.strftime('%b %d, %I:%M %p') if last_order else 'Never',
                'amount': total_amount,
                'phone': customer.mobile or '',
                'email': customer.email,
                'status': status,
                'avatar': initials
            })
        
        return Response({
            'status': 'success',
            'data': customer_data
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_recent_transactions(request):
    """Get recent transactions for retailer dashboard"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum, Count, Q
        from django.utils import timezone
        from datetime import timedelta
        
        # Get filter parameter
        view_mode = request.GET.get('mode', 'all')  # all, today, week
        
        # Base queryset
        orders = Order.objects.filter(
            retailer=user,
            status__in=['completed', 'pending', 'refunded']
        ).select_related('customer')
        
        # Apply filter
        today = timezone.now().date()
        
        if view_mode == 'today':
            orders = orders.filter(created_at__date=today)
        elif view_mode == 'week':
            last_week = today - timedelta(days=7)
            orders = orders.filter(created_at__date__gte=last_week)
        
        # Get recent orders (last 10)
        recent_orders = orders.order_by('-created_at')[:10]
        
        transactions = []
        for order in recent_orders:
            # Get item count
            item_count = order.items.aggregate(total=Sum('quantity'))['total'] or 0
            
            # Format time
            if order.created_at.date() == today:
                date_display = 'Today'
                time_display = order.created_at.strftime('%I:%M %p')
            elif order.created_at.date() == today - timedelta(days=1):
                date_display = 'Yesterday'
                time_display = order.created_at.strftime('%I:%M %p')
            else:
                date_display = order.created_at.strftime('%b %d')
                time_display = order.created_at.strftime('%I:%M %p')
            
            # Payment method mapping
            payment_method = order.payment_method or 'Cash'
            if payment_method == 'upi':
                payment_display = 'UPI'
            elif payment_method == 'card':
                payment_display = 'Card'
            elif payment_method == 'wallet':
                payment_display = 'Wallet'
            else:
                payment_display = 'Cash'
            
            transactions.append({
                'id': f"#TR-{order.id:04d}",
                'customer': order.customer.get_full_name() or order.customer.email.split('@')[0] if order.customer else 'Guest',
                'items': item_count,
                'amount': float(order.total_amount),
                'payment': payment_display,
                'status': order.status,
                'time': time_display,
                'date': date_display
            })
        
        return Response({
            'status': 'success',
            'data': transactions
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_low_stock_alerts(request):
    """Get low stock alerts for retailer dashboard"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Q, F
        from django.utils import timezone
        
        # Get filter parameter
        filter_type = request.GET.get('filter', 'all')  # all, critical, warning
        
        # Get products where stock <= threshold
        products = Product.objects.filter(
            seller=user,
            seller_type='retailer',
            status='active'
        ).select_related('category')
        
        # Filter by low stock condition
        low_stock_products = products.filter(stock__lte=F('threshold'))
        
        alerts = []
        for product in low_stock_products:
            # Determine status
            if product.stock == 0:
                status = 'critical'
            elif product.stock <= product.threshold // 2:
                status = 'critical'
            else:
                status = 'warning'
            
            # Apply filter
            if filter_type == 'critical' and status != 'critical':
                continue
            if filter_type == 'warning' and status != 'warning':
                continue
            
            # Get supplier from product brand or default
            supplier = product.brand or 'Generic Supplier'
            supplier_contact = '+91 00000 00000'  # Can be stored in Supplier model
            
            # Calculate lead time based on category or default
            lead_time = '3-5 days'
            if product.category:
                if product.category.name in ['Electronics', 'Gadgets']:
                    lead_time = '5-7 days'
                elif product.category.name in ['Furniture', 'Home Decor']:
                    lead_time = '7-10 days'
            
            alerts.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'currentStock': product.stock,
                'reorderLevel': product.threshold,
                'supplier': supplier,
                'supplierContact': supplier_contact,
                'leadTime': lead_time,
                'price': float(product.price),
                'status': status,
                'image': product.images.filter(is_primary=True).first().image.url if product.images.filter(is_primary=True).first() else '📦',
                'image_url': product.images.filter(is_primary=True).first().image.url if product.images.filter(is_primary=True).first() else None
            })
        
        # Sort by severity (critical first)
        alerts.sort(key=lambda x: 0 if x['status'] == 'critical' else 1)
        
        return Response({
            'status': 'success',
            'data': alerts
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_today_summary(request):
    """Get today's summary statistics for retailer dashboard"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum, Count
        from django.utils import timezone
        from datetime import datetime, timedelta
        
        today = timezone.now().date()
        start_of_day = datetime.combine(today, datetime.min.time())
        end_of_day = start_of_day + timedelta(days=1)
        
        # Get today's orders
        today_orders = Order.objects.filter(
            retailer=user,
            status='completed',
            created_at__gte=start_of_day,
            created_at__lt=end_of_day
        )
        
        # Total revenue
        total_revenue = today_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Total transactions
        total_transactions = today_orders.count()
        
        # Average bill
        average_bill = total_revenue / total_transactions if total_transactions > 0 else 0
        
        # Total items sold
        total_items = today_orders.aggregate(
            total=Sum('items__quantity')
        )['total'] or 0
        
        # Unique customers
        unique_customers = today_orders.values('customer').distinct().count()
        
        # Hourly breakdown
        hourly_breakdown = []
        busiest_hour = {'hour': '', 'transactions': 0, 'amount': 0}
        
        for hour in range(10, 22):  # 10 AM to 9 PM
            start_hour = start_of_day + timedelta(hours=hour)
            end_hour = start_hour + timedelta(hours=1)
            
            hour_orders = today_orders.filter(
                created_at__gte=start_hour,
                created_at__lt=end_hour
            )
            
            hour_transactions = hour_orders.count()
            hour_amount = hour_orders.aggregate(total=Sum('total_amount'))['total'] or 0
            
            hour_display = f"{hour} AM" if hour < 12 else f"{hour-12} PM" if hour > 12 else "12 PM"
            
            hourly_breakdown.append({
                'hour': hour_display,
                'transactions': hour_transactions,
                'amount': hour_amount
            })
            
            if hour_transactions > busiest_hour['transactions']:
                busiest_hour = {
                    'hour': f"{hour_display} - {(hour+1)} {hour_display.split()[1] if hour+1 != 12 else 'PM'}",
                    'transactions': hour_transactions,
                    'amount': hour_amount
                }
        
        # Payment methods breakdown
        payment_methods = {
            'upi': today_orders.filter(payment_method='upi').count(),
            'card': today_orders.filter(payment_method='card').count(),
            'cash': today_orders.filter(payment_method='cash').count(),
            'wallet': today_orders.filter(payment_method='wallet').count()
        }
        
        # Target (can be set from retailer settings or average of last 7 days)
        last_7_days = timezone.now().date() - timedelta(days=7)
        avg_daily_target = Order.objects.filter(
            retailer=user,
            status='completed',
            created_at__date__gte=last_7_days
        ).aggregate(avg=Sum('total_amount'))['avg'] or 25000
        avg_daily_target = avg_daily_target / 7 if avg_daily_target else 25000
        
        return Response({
            'status': 'success',
            'data': {
                'totalTransactions': total_transactions,
                'averageBill': round(average_bill, 2),
                'busiestHour': busiest_hour['hour'],
                'busiestHourSales': busiest_hour['amount'],
                'totalItems': total_items,
                'uniqueCustomers': unique_customers,
                'peakHourCustomers': busiest_hour['transactions'],
                'revenue': total_revenue,
                'target': round(avg_daily_target, 2),
                'paymentMethods': payment_methods,
                'hourlyBreakdown': hourly_breakdown
            }
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_quick_reorder(request):
    """Get quick reorder suggestions based on stock and sales velocity"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum, Count, F
        from django.utils import timezone
        from datetime import timedelta
        
        # Get last 30 days sales data
        last_30_days = timezone.now().date() - timedelta(days=30)
        next_7_days = timezone.now().date() + timedelta(days=7)
        
        # Get all retailer products
        products = Product.objects.filter(
            seller=user,
            seller_type='retailer',
            status='active'
        )
        
        # Calculate sales velocity for each product
        reorder_suggestions = []
        
        for product in products:
            # Get sales in last 30 days
            sales_data = OrderItem.objects.filter(
                product=product,
                order__retailer=user,
                order__status='completed',
                order__created_at__date__gte=last_30_days
            ).aggregate(
                total_sold=Sum('quantity'),
                order_count=Count('order', distinct=True)
            )
            
            total_sold = sales_data['total_sold'] or 0
            order_count = sales_data['order_count'] or 0
            
            # Calculate daily sales velocity
            daily_velocity = total_sold / 30 if total_sold > 0 else 0
            
            # Calculate days until out of stock
            if daily_velocity > 0:
                days_until_out = int(product.stock / daily_velocity) if product.stock > 0 else 0
            else:
                days_until_out = 999  # No sales, not urgent
            
            # Check if product needs reordering (stock below threshold or days until out < 7)
            needs_reorder = product.stock <= product.threshold or (days_until_out < 7 and days_until_out > 0)
            
            if not needs_reorder:
                continue
            
            # Determine urgency
            if product.stock == 0 or days_until_out <= 3:
                urgency = 'critical'
            elif product.stock <= product.threshold // 2 or days_until_out <= 7:
                urgency = 'high'
            elif product.stock <= product.threshold:
                urgency = 'medium'
            else:
                urgency = 'low'
            
            # Calculate suggested quantity (based on 30 days sales + safety stock)
            suggested_qty = int((daily_velocity * 30) + product.threshold) if daily_velocity > 0 else product.threshold * 2
            
            # Get supplier from brand or default
            supplier = product.brand or 'Generic Supplier'
            
            reorder_suggestions.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'currentStock': product.stock,
                'reorderLevel': product.threshold,
                'salesVelocity': round(daily_velocity, 1),
                'daysUntilOut': days_until_out,
                'suggestedQty': suggested_qty,
                'supplier': supplier,
                'price': float(product.price),
                'urgency': urgency,
                'image': product.images.filter(is_primary=True).first().image.url if product.images.filter(is_primary=True).first() else '📦',
                'image_url': product.images.filter(is_primary=True).first().image.url if product.images.filter(is_primary=True).first() else None
            })
        
        # Sort by urgency (critical first, then high, medium, low)
        urgency_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        reorder_suggestions.sort(key=lambda x: urgency_order.get(x['urgency'], 4))
        
        return Response({
            'status': 'success',
            'data': reorder_suggestions[:10]  # Limit to 10 suggestions
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)


#---------------------------------Retailers Reports------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_cogs(request):
    """
    Get Cost of Goods Sold for margin calculation
    """
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum, F, DecimalField, ExpressionWrapper
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        from commerce.models import OrderItem
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        product_id = request.GET.get('product_id')
        
        order_items = OrderItem.objects.filter(
            order__retailer=user,
            order__status='delivered'
        )
        
        if start_date:
            order_items = order_items.filter(order__created_at__date__gte=start_date)
        if end_date:
            order_items = order_items.filter(order__created_at__date__lte=end_date)
        if product_id:
            order_items = order_items.filter(product_id=product_id)
        
        # Calculate totals using .values() first to avoid aggregate error
        total_revenue = 0
        total_cogs = 0
        total_quantity = 0
        
        for item in order_items:
            total_revenue += float(item.quantity) * float(item.price)
            total_cogs += float(item.quantity) * float(item.product.cost_price) if item.product.cost_price else 0
            total_quantity += item.quantity
        
        gross_profit = total_revenue - total_cogs
        margin_percentage = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        return Response({
            'status': 'success',
            'data': {
                'summary': {
                    'total_revenue': round(total_revenue, 2),
                    'total_cogs': round(total_cogs, 2),
                    'gross_profit': round(gross_profit, 2),
                    'margin_percentage': round(margin_percentage, 2),
                    'total_quantity_sold': total_quantity
                },
                'filters': {
                    'start_date': start_date,
                    'end_date': end_date,
                    'product_id': product_id
                }
            }
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    
# ========== TAX REPORT ENDPOINTS ==========

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_tax_summary(request):
    """
    Get GST collected summary for retailer
    Query params: ?period=month&start_date=2026-01-01&end_date=2026-12-31
    """
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.db.models import Sum
        from django.utils import timezone
        from datetime import timedelta
        from commerce.models import Order
        
        # Get date range
        period = request.GET.get('period', 'month')  # day, week, month, quarter, year
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        today = timezone.now().date()
        
        if start_date and end_date:
            from datetime import datetime
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        elif period == 'day':
            start = today
            end = today
        elif period == 'week':
            start = today - timedelta(days=today.weekday())
            end = today
        elif period == 'month':
            start = today.replace(day=1)
            end = today
        elif period == 'quarter':
            quarter_month = ((today.month - 1) // 3) * 3 + 1
            start = today.replace(month=quarter_month, day=1)
            end = today
        elif period == 'year':
            start = today.replace(month=1, day=1)
            end = today
        else:
            start = today - timedelta(days=30)
            end = today
        
        # Get completed orders in date range
        orders = Order.objects.filter(
            retailer=user,
            status='delivered',
            created_at__date__gte=start,
            created_at__date__lte=end
        )
        
        # Calculate GST
        total_taxable_amount = orders.aggregate(total=Sum('total_amount'))['total'] or 0
        total_gst = orders.aggregate(total=Sum('tax_amount'))['total'] or 0
        
        # CGST and SGST are half of GST (assuming 50-50 split)
        cgst = float(total_gst) / 2
        sgst = float(total_gst) / 2
        
        # Get monthly breakdown for chart
        monthly_data = orders.annotate(
            month_date=TruncMonth('created_at')
        ).values('month_date').annotate(
            gst=Sum('tax_amount'),
            taxable=Sum('total_amount')
        ).order_by('month_date')
        
        monthly_breakdown = []
        for item in monthly_data:
            monthly_breakdown.append({
                'month': item['month_date'].strftime('%b %Y'),
                'gst': float(item['gst'] or 0),
                'taxable_amount': float(item['taxable'] or 0)
            })
        
        return Response({
            'status': 'success',
            'data': {
                'summary': {
                    'period': {
                        'start_date': start.isoformat(),
                        'end_date': end.isoformat(),
                        'period_type': period
                    },
                    'total_taxable_amount': float(total_taxable_amount),
                    'total_gst_collected': float(total_gst),
                    'cgst': round(cgst, 2),
                    'sgst': round(sgst, 2),
                    'total_orders': orders.count()
                },
                'monthly_breakdown': monthly_breakdown
            }
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retailer_gst_returns(request):
    """
    Get list of GST returns filed/pending
    """
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from .models import GSTReturn
        
        # Get query parameters
        status_filter = request.GET.get('status')  # filed, pending, all
        year = request.GET.get('year')
        
        returns = GSTReturn.objects.filter(retailer=user)
        
        if status_filter and status_filter != 'all':
            returns = returns.filter(status=status_filter)
        if year:
            returns = returns.filter(period__year=year)
        
        returns = returns.order_by('-period')
        
        # Pagination
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        start = (page - 1) * per_page
        end = start + per_page
        
        total = returns.count()
        paginated = returns[start:end]
        
        from .serializers import GSTReturnSerializer
        serializer = GSTReturnSerializer(paginated, many=True)
        
        # Summary stats
        filed_count = returns.filter(status='filed').count()
        pending_count = returns.filter(status='pending').count()
        total_tax = returns.filter(status='filed').aggregate(total=Sum('tax_amount'))['total'] or 0
        
        return Response({
            'status': 'success',
            'data': {
                'returns': serializer.data,
                'summary': {
                    'filed': filed_count,
                    'pending': pending_count,
                    'total_tax_paid': float(total_tax)
                },
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': total,
                    'total_pages': (total + per_page - 1) // per_page
                }
            }
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def file_gst_return(request):
    """
    File a new GST return
    """
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from .models import GSTReturn
        from .serializers import FileGSTReturnSerializer
        from django.utils import timezone
        from commerce.models import Order
        from django.db.models import Sum
        
        serializer = FileGSTReturnSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({'status': 'error', 'errors': serializer.errors}, status=400)
        
        data = serializer.validated_data
        period = data['period']
        
        # Parse period (e.g., "Q1 2026" or "Jan-Mar 2026")
        from datetime import datetime
        
        # Calculate tax for the period
        # For quarterly returns
        if 'Q' in period:
            quarter_num = int(period.split('Q')[1].split()[0])
            year = int(period.split()[-1])
            
            if quarter_num == 1:
                start_date = datetime(year, 1, 1).date()
                end_date = datetime(year, 3, 31).date()
            elif quarter_num == 2:
                start_date = datetime(year, 4, 1).date()
                end_date = datetime(year, 6, 30).date()
            elif quarter_num == 3:
                start_date = datetime(year, 7, 1).date()
                end_date = datetime(year, 9, 30).date()
            else:
                start_date = datetime(year, 10, 1).date()
                end_date = datetime(year, 12, 31).date()
        else:
            # Monthly period
            return Response({'status': 'error', 'message': 'Invalid period format'}, status=400)
        
        # Check if return already exists for this period
        if GSTReturn.objects.filter(retailer=user, period=period).exists():
            return Response({
                'status': 'error',
                'message': f'GST return for period {period} already filed'
            }, status=400)
        
        # Calculate tax from orders in period
        orders = Order.objects.filter(
            retailer=user,
            status='delivered',
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        tax_amount = orders.aggregate(total=Sum('tax_amount'))['total'] or 0
        
        # Create GST return
        gst_return = GSTReturn.objects.create(
            retailer=user,
            period=period,
            tax_amount=tax_amount,
            status='filed',
            filed_date=timezone.now().date(),
            due_date=data.get('due_date', timezone.now().date() + timedelta(days=30))
        )
        
        from .serializers import GSTReturnSerializer
        response_serializer = GSTReturnSerializer(gst_return)
        
        return Response({
            'status': 'success',
            'message': f'GST return for period {period} filed successfully',
            'data': response_serializer.data
        }, status=201)
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def retailer_export_report(request):
    """Generate PDF/Excel report for retailer"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        import io
        import xlsxwriter
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from commerce.models import Order
        from datetime import datetime
        
        report_type = request.data.get('report_type', 'sales')  # sales, products, customers, tax, profit
        format_type = request.data.get('format', 'excel')  # pdf, excel
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        
        # Get orders data
        orders = Order.objects.filter(
            retailer=user,
            status='delivered',
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        if format_type == 'excel':
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet(report_type.title())
            
            # Headers
            headers = ['Order ID', 'Date', 'Customer', 'Amount', 'Status', 'Items']
            for col, header in enumerate(headers):
                worksheet.write(0, col, header)
            
            # Data
            for row, order in enumerate(orders, start=1):
                worksheet.write(row, 0, order.order_number)
                worksheet.write(row, 1, order.created_at.strftime('%Y-%m-%d'))
                worksheet.write(row, 2, order.customer.email)
                worksheet.write(row, 3, float(order.grand_total))
                worksheet.write(row, 4, order.status)
                worksheet.write(row, 5, order.items.count())
            
            workbook.close()
            output.seek(0)
            
            filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response
            
        else:  # PDF
            response = HttpResponse(content_type='application/pdf')
            filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            doc = SimpleDocTemplate(response, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []
            
            # Title
            title = Paragraph(f"{report_type.title()} Report", styles['Title'])
            elements.append(title)
            
            # Table data
            data = [['Order ID', 'Date', 'Amount']]
            for order in orders[:50]:
                data.append([order.order_number, order.created_at.strftime('%Y-%m-%d'), f"₹{order.grand_total}"])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            doc.build(elements)
            return response
            
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def retailer_email_report(request):
    """Send report via email"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    try:
        from django.core.mail import EmailMessage
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        import io
        import xlsxwriter
        from datetime import datetime
        
        email_to = request.data.get('email')
        report_type = request.data.get('report_type', 'sales')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        
        if not email_to:
            return Response({'status': 'error', 'message': 'Email address required'}, status=400)
        
        # Generate Excel file
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(report_type.title())
        
        from commerce.models import Order
        orders = Order.objects.filter(
            retailer=user,
            status='delivered',
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        headers = ['Order ID', 'Date', 'Customer', 'Amount', 'Status']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
        
        for row, order in enumerate(orders, start=1):
            worksheet.write(row, 0, order.order_number)
            worksheet.write(row, 1, order.created_at.strftime('%Y-%m-%d'))
            worksheet.write(row, 2, order.customer.email)
            worksheet.write(row, 3, float(order.grand_total))
            worksheet.write(row, 4, order.status)
        
        workbook.close()
        output.seek(0)
        
        # Send email
        subject = f"{report_type.title()} Report - {datetime.now().strftime('%Y-%m-%d')}"
        html_message = render_to_string('email/report_email.html', {
            'user_name': user.get_full_name() or user.email,
            'report_type': report_type,
            'start_date': start_date,
            'end_date': end_date,
            'total_orders': orders.count(),
            'total_amount': orders.aggregate(total=Sum('grand_total'))['total'] or 0
        })
        plain_message = strip_tags(html_message)
        
        email = EmailMessage(
            subject, plain_message, None, [email_to]
        )
        email.attach(f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.xlsx", output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.send()
        
        return Response({
            'status': 'success',
            'message': f'Report sent to {email_to}'
        })
        
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)
    

# ========== SCHEDULED REPORTS ==========

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_scheduled_reports(request):
    """List all scheduled reports for retailer"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    from .models import ScheduledReport
    from .serializers import ScheduledReportSerializer
    
    reports = ScheduledReport.objects.filter(retailer=user).order_by('-created_at')
    serializer = ScheduledReportSerializer(reports, many=True)
    
    return Response({'status': 'success', 'data': serializer.data})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_scheduled_report(request):
    """Create a new scheduled report"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    from .models import ScheduledReport
    from .serializers import ScheduledReportCreateSerializer
    
    serializer = ScheduledReportCreateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({'status': 'error', 'errors': serializer.errors}, status=400)
    
    data = serializer.validated_data
    
    report = ScheduledReport.objects.create(
        retailer=user,
        name=data['name'],
        report_type=data['report_type'],
        frequency=data['frequency'],
        format_type=data.get('format_type', 'excel'),
        recipients=data['recipients'],
        is_active=data.get('is_active', True)
    )
    
    from .serializers import ScheduledReportSerializer
    response_serializer = ScheduledReportSerializer(report)
    
    return Response({
        'status': 'success',
        'message': 'Scheduled report created successfully',
        'data': response_serializer.data
    }, status=201)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_scheduled_report(request, report_id):
    """Update a scheduled report"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    from .models import ScheduledReport
    from .serializers import ScheduledReportUpdateSerializer
    
    try:
        report = ScheduledReport.objects.get(id=report_id, retailer=user)
    except ScheduledReport.DoesNotExist:
        return Response({'status': 'error', 'message': 'Scheduled report not found'}, status=404)
    
    serializer = ScheduledReportUpdateSerializer(data=request.data, partial=True)
    if not serializer.is_valid():
        return Response({'status': 'error', 'errors': serializer.errors}, status=400)
    
    data = serializer.validated_data
    
    if 'name' in data:
        report.name = data['name']
    if 'frequency' in data:
        report.frequency = data['frequency']
    if 'format_type' in data:
        report.format_type = data['format_type']
    if 'recipients' in data:
        report.recipients = data['recipients']
    if 'is_active' in data:
        report.is_active = data['is_active']
    
    report.save()
    
    from .serializers import ScheduledReportSerializer
    response_serializer = ScheduledReportSerializer(report)
    
    return Response({
        'status': 'success',
        'message': 'Scheduled report updated successfully',
        'data': response_serializer.data
    })


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_scheduled_report(request, report_id):
    """Delete a scheduled report"""
    user = request.user
    
    if user.role != 'retailer':
        return Response({'status': 'error', 'message': 'Only retailers can access'}, status=403)
    
    from .models import ScheduledReport
    
    try:
        report = ScheduledReport.objects.get(id=report_id, retailer=user)
    except ScheduledReport.DoesNotExist:
        return Response({'status': 'error', 'message': 'Scheduled report not found'}, status=404)
    
    report.delete()
    
    return Response({
        'status': 'success',
        'message': 'Scheduled report deleted successfully'
    })