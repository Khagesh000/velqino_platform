from django.db import transaction
from django.core.paginator import Paginator
from django.core.cache import cache
from ..models import Product, Category, ProductImage
from ..utils.product_helpers import ProductHelpers
from ..tasks import process_bulk_images_task, process_bulk_video_task
import logging
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

logger = logging.getLogger(__name__)


class ProductService:
    """Business logic layer with connection pooling and indexing"""
    
    @staticmethod
    def create_product(seller, data, front_image=None, back_image=None):
        """Create single product with images"""
        with transaction.atomic():
            product = Product.objects.create(
                seller=seller,
                name=data.get('name'),
                price=data['price'],
                cost=data['cost'],
                category_id=data.get('category_id'),
                brand=data.get('brand', ''),
                description=data.get('description', ''),
                pattern=data.get('pattern', ''),
                primary_color=data.get('primary_color', ''),
                status='active'
            )
            
            # Add front image
            if front_image:
                ProductImage.objects.create(
                    product=product,
                    image=front_image,
                    is_primary=True,
                    is_front=True,
                    order=0
                )
            
            # Add back image
            if back_image:
                ProductImage.objects.create(
                    product=product,
                    image=back_image,
                    is_primary=False,
                    is_front=False,
                    order=1
                )
            
            ProductHelpers.invalidate_product_caches(seller.id)
            return product
    
    @staticmethod
    def get_products_with_filters(seller_id, filters, page=1, per_page=20):
        """Get products with filtering, pagination, and indexing"""
        cache_key = f"product:list:{seller_id}:{hash(str(filters))}:{page}:{per_page}"
        cached = cache.get(cache_key)
        
        if cached:
            return cached
        
        queryset = Product.objects.filter(seller_id=seller_id).select_related('category')
        
        # Apply filters with indexes
        if filters.get('category'):
            queryset = queryset.filter(category_id=filters['category'])
        
        if filters.get('status'):
            queryset = queryset.filter(status=filters['status'])
        
        if filters.get('min_price'):
            queryset = queryset.filter(price__gte=filters['min_price'])
        
        if filters.get('max_price'):
            queryset = queryset.filter(price__lte=filters['max_price'])
        
        if filters.get('pattern'):
            queryset = queryset.filter(pattern=filters['pattern'])
        
        if filters.get('low_stock'):
            queryset = ProductHelpers.get_low_stock_products(seller_id)
        
        if filters.get('search'):
            queryset = ProductHelpers.search_products(filters['search'], seller_id)
        
        # Paginate
        paginator = Paginator(queryset, per_page)
        products = paginator.get_page(page)
        
        result = {
            'products': products,
            'total': paginator.count,
            'page': page,
            'per_page': per_page,
            'total_pages': paginator.num_pages
        }
        
        cache.set(cache_key, result, 300)  # 5 minutes
        return result
    
    @staticmethod
    def bulk_operation(seller_id, product_ids, operation, value=None):
        """Perform bulk operations"""
        products = Product.objects.filter(seller_id=seller_id, id__in=product_ids)
        
        if operation == 'delete':
            count = products.delete()[0]
            ProductHelpers.invalidate_product_caches(seller_id)
            return {'deleted': count}
        
        elif operation == 'update_status':
            count = products.update(status=value)
            ProductHelpers.invalidate_product_caches(seller_id)
            return {'updated': count}
        
        elif operation == 'update_price':
            with transaction.atomic():
                for product in products:
                    product.price = value
                    product.save()
            ProductHelpers.invalidate_product_caches(seller_id)
            return {'updated': len(product_ids)}
        
        return {'error': 'Invalid operation'}
    
    @staticmethod
    def trigger_bulk_image_processing(seller_id, images, common_data):
        """Trigger async task for bulk image processing"""
        task = process_bulk_images_task.delay(
            seller_id=seller_id,
            images_data=[img.read() for img in images],
            common_price=common_data['common_price'],
            common_cost=common_data['common_cost'],
            category_id=common_data.get('category_id'),
            common_name_prefix=common_data.get('common_name_prefix', 'Product'),
            brand=common_data.get('brand', ''),
            description=common_data.get('description', ''),
            upload_mode=common_data.get('upload_mode', 'front_back'),
            sizes=common_data.get('sizes', []),
        )
        return {'task_id': task.id, 'status': 'queued'}
    
    @staticmethod
    def trigger_bulk_video_processing(seller_id, video, common_data):
        """Trigger async task for bulk video processing"""
        task = process_bulk_video_task.delay(
            seller_id=seller_id,
            video_data=video.read(),
            number_of_products=common_data['number_of_products'],
            common_price=common_data['common_price'],
            common_cost=common_data['common_cost'],
            category_id=common_data.get('category_id'),
            common_name_prefix=common_data.get('common_name_prefix', 'Product'),
            brand=common_data.get('brand', ''),
            description=common_data.get('description', ''),
            sizes=common_data.get('sizes', []),
            grid_rows=common_data.get('grid_rows', 2),
            grid_columns=common_data.get('grid_columns', 5)
        )
        return {'task_id': task.id, 'status': 'queued'}



    # ADD this method to existing ProductService class

    @staticmethod
    def create_single_product_advanced(seller_id, data, task_id=None):
        """Advanced single product creation with WebSocket progress"""
        from channels.layers import get_channel_layer
        from django.core.cache import cache
        
        channel_layer = get_channel_layer()
        room_group_name = f"product_task_{task_id}" if task_id else f"product_{seller_id}"
        
        # Send progress
        from asyncio import new_event_loop, set_event_loop
        loop = new_event_loop()
        set_event_loop(loop)
        
        async def send_progress(progress, message):
            await channel_layer.group_send(
                room_group_name,
                {'type': 'send_progress', 'data': {'progress': progress, 'message': message}}
            )
        
        loop.run_until_complete(send_progress(10, "Creating product record..."))
        
        # Create product
        from ..models import Product
        import uuid
        from django.utils.text import slugify
        
        sku = data.get('sku') or f"PROD-{uuid.uuid4().hex[:8].upper()}"
        product = Product.objects.create(
            seller_id=seller_id,
            name=data['name'],
            sku=sku,
            slug=f"{slugify(data['name'])}-{sku}"[:50],
            category_id=data.get('category_id'),
            brand=data.get('brand', ''),
            description=data.get('description', ''),
            price=data['price'],
            cost=data.get('cost'),
            compare_price=data.get('compare_price'),
            stock=data.get('stock', 0),
            threshold=data.get('threshold', 10),
            status=data.get('status', 'draft'),
            weight=data.get('weight')
        )
        
        loop.run_until_complete(send_progress(50, "Caching product..."))
        
        # Cache
        cache_key = f"product:{product.id}"
        cache.set(cache_key, product, 300)
        
        # Invalidate caches
        from ..utils.product_helpers import ProductHelpers
        ProductHelpers.invalidate_product_caches(seller_id)
        
        loop.run_until_complete(send_progress(100, "Product created successfully!"))
        loop.close()
        
        return {'product_id': product.id, 'sku': product.sku}
    


    @staticmethod
    def create_product_with_variants(seller, data, sizes=None, images=None):
        from ..models import Product, ProductVariant, ProductImage
        import uuid
        from django.utils.text import slugify
        
        sku = data.get('sku') or f"PROD-{uuid.uuid4().hex[:8].upper()}"
        
        product = Product.objects.create(
            seller=seller,
            name=data['name'],
            sku=sku,
            slug=f"{slugify(data['name'])}-{sku}"[:50],
            category_id=data.get('category_id'),
            brand=data.get('brand', ''),
            description=data.get('description', ''),
            price=data['price'],
            cost=data.get('cost'),
            compare_price=data.get('compare_price'),
            stock=data.get('stock', 0),
            threshold=data.get('threshold', 10),
            status=data.get('status', 'draft'),
            weight=data.get('weight')
        )
        
        # ✅ Create variants from sizes parameter (FormData)
        if sizes:
            for size in sizes:
                ProductVariant.objects.create(
                    product=product,
                    size=size,
                    sku=f"{sku}-{size}",
                    stock=product.stock,
                    price=product.price
                )
        
        # ✅ Create images from FILES
        if images:
            for idx, img in enumerate(images):
                ProductImage.objects.create(
                    product=product,
                    image=img,
                    is_primary=(idx == 0),
                    order=idx
                )
        
        return product
    

    @staticmethod
    def export_products_to_csv(seller_id):
        """Export products to CSV with caching"""
        cache_key = f"export:csv:{seller_id}"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return cached_data
        
        products = Product.objects.filter(seller_id=seller_id).select_related('category').prefetch_related('variants', 'images')
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            'ID', 'Name', 'SKU', 'Price', 'Cost', 'Compare Price', 'Stock', 
            'Threshold', 'Category', 'Brand', 'Description', 'Status', 
            'Pattern', 'Primary Color', 'Weight', 'Created At', 'Image URLs', 'Variants'
        ])
        
        # Data rows
        for product in products:
            images_urls = ', '.join([img.image.url for img in product.images.all()[:5]])
            variants_info = ', '.join([f"{v.size}" for v in product.variants.all()])
            
            writer.writerow([
                product.id, product.name, product.sku, product.price, product.cost,
                product.compare_price or '', product.stock, product.threshold,
                product.category.name if product.category else '', product.brand,
                product.description[:200], product.status, product.pattern,
                product.primary_color, product.weight or '', product.created_at,
                images_urls, variants_info
            ])
        
        result = output.getvalue()
        cache.set(cache_key, result, 300)  # Cache for 5 minutes
        return result
    
    @staticmethod
    def export_products_to_excel(seller_id):
        """Export products to Excel with formatting"""
        cache_key = f"export:excel:{seller_id}"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return cached_data
        
        products = Product.objects.filter(seller_id=seller_id).select_related('category').prefetch_related('variants', 'images')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Headers with styling
        headers = ['ID', 'Name', 'SKU', 'Price', 'Cost', 'Compare Price', 'Stock', 
                   'Threshold', 'Category', 'Brand', 'Description', 'Status', 
                   'Pattern', 'Primary Color', 'Weight', 'Created At', 'Images', 'Variants']
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row_idx, product in enumerate(products, 2):
            images_urls = ', '.join([img.image.url for img in product.images.all()[:3]])
            variants_info = ', '.join([f"{v.size}:{v.stock}" for v in product.variants.all()])
            
            ws.cell(row=row_idx, column=1, value=product.id)
            ws.cell(row=row_idx, column=2, value=product.name)
            ws.cell(row=row_idx, column=3, value=product.sku)
            ws.cell(row=row_idx, column=4, value=float(product.price))
            ws.cell(row=row_idx, column=5, value=float(product.cost) if product.cost else 0)
            ws.cell(row=row_idx, column=6, value=float(product.compare_price) if product.compare_price else '')
            ws.cell(row=row_idx, column=7, value=product.stock)
            ws.cell(row=row_idx, column=8, value=product.threshold)
            ws.cell(row=row_idx, column=9, value=product.category.name if product.category else '')
            ws.cell(row=row_idx, column=10, value=product.brand)
            ws.cell(row=row_idx, column=11, value=product.description[:200])
            ws.cell(row=row_idx, column=12, value=product.status)
            ws.cell(row=row_idx, column=13, value=product.pattern)
            ws.cell(row=row_idx, column=14, value=product.primary_color)
            ws.cell(row=row_idx, column=15, value=float(product.weight) if product.weight else '')
            ws.cell(row=row_idx, column=16, value=product.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_idx, column=17, value=images_urls)
            ws.cell(row=row_idx, column=18, value=variants_info)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        result = output.getvalue()
        cache.set(cache_key, result, 300)
        return result
    
    @staticmethod
    def invalidate_export_cache(seller_id):
        """Clear export cache when products change"""
        cache.delete(f"export:csv:{seller_id}")
        cache.delete(f"export:excel:{seller_id}")